import nibabel as nib
import numpy as np
from constants import Constants
from werkzeug.datastructures import MultiDict
import scipy.ndimage as ndimage
import os, sys
import tempfile
from scipy.ndimage import label
import pathlib
from openpyxl import load_workbook

def get_panTS_id(index):
    cur_case_id = str(index)
    iter = max(0, 8 - len(str(index)))
    for _ in range(iter):
        cur_case_id = "0" + cur_case_id
    cur_case_id = "PanTS_" + cur_case_id    
    return cur_case_id

def has_large_connected_component(slice_mask, threshold=8):
    """
    Check if there is a connected component larger than a threshold in a 2D mask.
    """
    labeled, num_features = label(slice_mask)
    sizes = np.bincount(labeled.ravel())
    sizes[0] = 0  # ignore background
    return np.any(sizes > threshold)


class NpzProcessor:
    def __init__(self, main_npz_path=None, clabel_path=None, organ_intensities=None):
        self._main_nifti_path = main_npz_path
        self._clabel_path = clabel_path
        self.number_max = 999999
        self._organ_intensities = organ_intensities
    
    def set_organ_intensities(self, organ_intensities):
        self._organ_intensities = organ_intensities

    @classmethod
    def from_clabel_path(cls, clabel_path):
        
        return cls(None, clabel_path)
    
    # not used
    def calculate_mean_hu_with_erosion(self, binary_mask, ct_array):
        """
        Calculate mean HU using erosion to avoid edge noise.
        """
        erosion_array = ndimage.binary_erosion(binary_mask, structure=Constants.STRUCTURING_ELEMENT)
        hu_values = ct_array[erosion_array > 0]

        if hu_values.size == 0:
            hu_values = ct_array[binary_mask > 0]

        if hu_values.size == 0:
            return 0

        return round(float(np.mean(hu_values)), Constants.DECIMAL_PRECISION_HU)
    
    def npz_to_nifti(self, id: int, combined_label=True, save=True):
        subfolder = "LabelTr" if combined_label else "ImageTr"
        if id >= 9000:
            subfolder = "LabelTe" if combined_label else "ImageTe"
        
        if combined_label:    
            dir_path = pathlib.Path(f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(id)}/{Constants.COMBINED_LABELS_FILENAME}")
        else:
            dir_path = pathlib.Path(f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(id)}/ct.npz")
        
        metadata_path = pathlib.Path(f"{Constants.PANTS_PATH}/data/metadata.xlsx")
        
        wb = load_workbook(metadata_path)
        sheet = wb["PanTS_metadata"]
        spacing = []
        # ! col_0 - id, col_1 - shape, col_2 - spacing, col_3  - ct_phase, col_4 - sex, col_5 - age
        for row in sheet.iter_rows(values_only=True):
            if row[0] == get_panTS_id(id):
                # make into int
                spacing = row[2].strip("()").split(", ")
                break
                            
        if len(spacing) == 0:
            f"Could not find metadata for PanTS ID: {get_panTS_id(id)}"
            return None

        for i in range(len(spacing)):
            spacing[i] = float(spacing[i])

        arr = np.load(dir_path)["data"].astype(np.float32)
        affine = np.diag(spacing + [1])
        img = nib.nifti1.Nifti1Image(arr, affine=affine)
        img.header.set_zooms(spacing) 

        nib.save(img, dir_path.with_suffix(".nii.gz"))
            
            
    def combine_labels(self, id: int, save=True):   
        """
        Merge multiple label masks into one combined segmentation and re-index the labels.
        """
        organ_intensities = {}
        segment_subfolder = "LabelTr"
        if id >= 9000:
            segment_subfolder = "LabelTe"   
        
        dir_path = pathlib.Path(f"{Constants.PANTS_PATH}/data/{segment_subfolder}/{get_panTS_id(id)}/segmentations")
        npz_files = list(dir_path.glob("*.npz"))
        
        combined_labels_img_data = None

        for i in range(len(npz_files)):
            filename = npz_files[i].name
            data = np.load(dir_path /filename)["data"]
            # not contiguous, may need original ct to get shape
            if combined_labels_img_data is None:
                combined_labels_img_data = np.ndarray(shape=data.shape, dtype=np.float64)

            scaled = data * np.float64(i + 1)
            combined_labels_img_data = np.maximum(combined_labels_img_data, scaled)

            organ_intensities[filename] = i + 1

        if save:
            save_path = f"{Constants.PANTS_PATH}/data/{segment_subfolder}/{get_panTS_id(id)}/{Constants.COMBINED_LABELS_FILENAME}"
            np.savez_compressed(save_path, data=combined_labels_img_data)

        return combined_labels_img_data, organ_intensities

    def __str__(self):
        return f"Npz Processor Object\n"
