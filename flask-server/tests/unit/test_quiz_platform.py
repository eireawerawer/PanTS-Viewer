from __future__ import annotations

import copy
import gzip
import io
import json
import tarfile
import threading
import zipfile
from pathlib import Path

import nibabel as nib
import numpy as np
import openpyxl
import pytest
from flask import Flask
from PIL import Image

import api.education as education_api
from scripts.stage_quiz_dataset import extract_combined_labels
from services.live_quiz import (
    QuizPackError,
    QuizPackRegistry,
    get_pack,
    public_pack,
    public_question,
    reveal_for,
    validate_pack,
)
from services.live_room_store import LiveRoomError, LiveRoomStore
from services.quiz_generation import (
    Candidate,
    LabelProfileRegistry,
    QuizGenerationError,
    TARGETS,
    generate_candidate,
    generate_release,
    discover_candidates,
    GeneratedCandidate,
    _mark_outliers,
    _primary_component,
    _review_sample,
    _foreground_voxel_lps,
    maximum_axial_diameter,
)
from services.quiz_practice_store import QuizPracticeError, QuizPracticeStore
from services.quiz_telemetry import QuizTelemetry


def catalog_with(count: int = 10) -> dict:
    base = get_pack("radworld-case-35-v1")
    packs = []
    for index in range(count):
        pack = copy.deepcopy(base)
        positive = index < 7
        pack["pack_id"] = f"test-pack-{index}-v1"
        pack["case_id"] = str(100 + index)
        pack["difficulty"] = "easy"
        pack["approval"] = {"status": "approved", "reviewer": "test-fixture"}
        pack["provenance"].update({
            "dataset_id": f"PanTS_{100 + index:08d}",
            "template_version": "pancreas-imaging-chain/1.1.0",
            "label_profile_id": "case35-reviewed-v1",
        })
        pack["ground_truth"]["lesion_present"] = positive
        if not positive:
            pack["ground_truth"].update({
                "lesion_labels": [],
                "lesion_centroid_lps": None,
                "pancreatic_region": None,
                "max_axial_diameter_mm": None,
                "reference_measurement_lps": None,
                "reveal_mask": {"kind": "labels", "source_labels": [], "output_label": 1},
            })
            pack["questions"][1]["correct_choice_id"] = "no"
            pack["questions"][2]["correct_choice_id"] = "not_applicable"
            pack["questions"][3]["correct_choice_id"] = "no_abnormality"
            pack["questions"][3]["reveal_viewer_cue"] = {
                "show_lesion_overlay": False,
                "crosshair_lps": pack["ground_truth"]["crosshair_lps"],
            }
        packs.append(validate_pack(pack))
    return {"catalog_id": "test-v1", "catalog_version": 1, "packs": packs}


def v2_pack(base: dict, *, with_report_question: bool) -> dict:
    pack = copy.deepcopy(base)
    pack["pack_id"] = pack["pack_id"].removesuffix("-v1") + "-v2"
    pack["version"] = 2
    pack["generator_version"] = "bodymaps-vqa-generator/2.0.0"
    pack["validator_version"] = "bodymaps-vqa-validator/2.0.0"
    pack["approval"] = {"status": "approved", "reviewer": "test-fixture"}
    if with_report_question:
        pack["questions"].insert(3, {
            "id": "report_finding",
            "prompt": "How is pancreatic-tail lesion described relative to pancreas?",
            "choices": [
                {"id": "hyperattenuating", "label": "hyperattenuating", "claims": {}},
                {"id": "isoattenuating", "label": "isoattenuating", "claims": {}},
                {"id": "hypoattenuating", "label": "hypoattenuating", "claims": {}},
                {"id": "not_described", "label": "not described", "claims": {}},
            ],
            "correct_choice_id": "hyperattenuating",
            "explanation": "Structured report describes lesion as hyperattenuating relative to pancreas.",
            "source_label": "Structured report finding",
            "viewer_cue": copy.deepcopy(pack["questions"][0]["viewer_cue"]),
        })
        pack["report_ground_truth"] = {
            "report_sha256": "b" * 64,
            "selected_field": "lesion_attenuation",
            "evidence_sentence": "Pancreas lesion 1. Location: pancreas tail.",
            "parser_version": "radgpt-structured-report/1.0.0",
            "validation_outcome": "matched_report_lesion_to_mask",
        }
    else:
        pack["report_ground_truth"] = {
            "report_sha256": None,
            "selected_field": None,
            "evidence_sentence": None,
            "parser_version": "radgpt-structured-report/1.0.0",
            "validation_outcome": "report_missing",
        }
    return validate_pack(pack)


def write_case(root: Path, case_id: int, *, affine=None, mask_affine=None, lesion=True, multiple=False):
    dataset_id = f"PanTS_{case_id:08d}"
    image_dir = root / "image_only" / dataset_id
    mask_dir = root / "mask_only" / dataset_id
    image_dir.mkdir(parents=True)
    mask_dir.mkdir(parents=True)
    affine = np.eye(4) if affine is None else affine
    mask_affine = affine if mask_affine is None else mask_affine
    ct = np.zeros((16, 16, 5), dtype=np.int16)
    mask = np.zeros_like(ct, dtype=np.uint8)
    mask[2:5, 3:13, 2] = 17
    mask[2:5, 3:6, 2] = 19
    mask[2:5, 6:9, 2] = 18
    mask[2:5, 10:13, 2] = 20
    if lesion:
        mask[3:5, 10:13, 2] = 28
        if multiple:
            mask[8:10, 10:12, 3] = 28
    nib.save(nib.Nifti1Image(ct, affine), image_dir / "ct.nii.gz")
    nib.save(nib.Nifti1Image(mask, mask_affine), mask_dir / "combined_labels.nii.gz")


def profile_for(*case_ids: int) -> LabelProfileRegistry:
    return LabelProfileRegistry({
        "profile_version": 1,
        "profiles": [{
            "profile_id": "test-profile",
            "match": {"case_ids": list(case_ids)},
            "labels": {
                "pancreas": [17, 18, 19, 20],
                "head": [19],
                "body": [18],
                "tail": [20],
                "lesion": [28],
            },
            "reveal_output_label": 1,
            "mesh_organ_id": 22,
        }],
    })


def synthetic_structured_report(*, lesion: bool) -> str:
    lesion_text = (
        "Pancreas lesion 1:\nLocation: pancreas tail.\n"
        "Size: 0.2 x 0.1 cm (image 2). Volume: 0.1 cm^3.\n"
        "Enhancement relative to pancreas: Hyperattenuating (HU value is 80+/-5).\n"
        if lesion else ""
    )
    return (
        "FINDINGS:\n"
        "Spleen:\nNormal size (volume: 200.0 cm^3).\nMean HU value: 45.0 +/- 5.0.\n"
        "Pancreas:\nNormal size (volume: 70.0 cm^3).\nMean HU value: 50.0 +/- 6.0.\n"
        f"{lesion_text}"
        "Kidney:\nNormal size (total kidney volume: 300.0 cm^3).\n"
        "Liver:\nNormal size (volume: 1400.0 cm^3).\n"
        "IMPRESSION:\nNo additional report finding."
    )


def test_public_pack_and_playlist_payloads_do_not_leak_server_truth():
    pack = get_pack("radworld-case-35-v1")
    public = public_pack(pack)
    encoded = json.dumps(public)
    assert "ground_truth" not in public
    assert "correct_choice_id" not in encoded
    assert "explanation" not in encoded
    assert "reveal_mask" not in encoded
    assert "claims" not in encoded
    assert "tags" not in public
    assert "tumor-positive" not in encoded
    playlists = QuizPackRegistry(catalog_with()).public_playlists()
    assert all("positive_count" not in item and "normal_count" not in item and "target_mix" not in item for item in playlists)


def test_public_choice_order_is_seeded_and_keeps_choice_ids():
    pack = get_pack("radworld-case-35-v1")
    first = public_question(pack, 0, choice_seed="attempt-a")
    repeated = public_question(pack, 0, choice_seed="attempt-a")
    second = public_question(pack, 0, choice_seed="attempt-b")
    canonical_ids = {choice["id"] for choice in pack["questions"][0]["choices"]}
    assert first == repeated
    assert {choice["id"] for choice in first["choices"]} == canonical_ids
    assert {choice["id"] for choice in second["choices"]} == canonical_ids
    assert first["choices"] != second["choices"]


def test_geometry_is_orientation_safe_component_aware_and_uses_foreground_cue():
    binary = np.zeros((12, 10, 6), dtype=bool)
    binary[2:8, 2:7, 3] = True
    binary[10:12, 8:10, 3] = True
    primary, count = _primary_component(binary)
    assert count == 2

    canonical = np.diag([0.7, 1.3, 4.0, 1.0])
    diameter, endpoints, _ = maximum_axial_diameter(primary, canonical)
    transform = nib.orientations.ornt_transform(
        nib.orientations.axcodes2ornt(("R", "A", "S")),
        nib.orientations.axcodes2ornt(("I", "P", "L")),
    )
    permuted = nib.orientations.apply_orientation(primary, transform)
    permuted_affine = canonical @ nib.orientations.inv_ornt_aff(transform, primary.shape)
    permuted_diameter, permuted_endpoints, _ = maximum_axial_diameter(permuted, permuted_affine)
    assert permuted_diameter == pytest.approx(diameter)
    assert sorted(map(tuple, permuted_endpoints)) == pytest.approx(sorted(map(tuple, endpoints)))

    ring = np.zeros((9, 9, 3), dtype=bool)
    ring[1:8, 1:8, 1] = True
    ring[3:6, 3:6, 1] = False
    cue_lps = _foreground_voxel_lps(ring, np.eye(4))
    cue_ras = np.asarray(cue_lps)
    cue_ras[:2] *= -1
    cue_ijk = np.rint(cue_ras).astype(int)
    assert ring[tuple(cue_ijk)]


def test_pack_validation_rejects_non_finite_ground_truth():
    pack = copy.deepcopy(get_pack("radworld-case-35-v1"))
    pack["ground_truth"]["crosshair_lps"][0] = float("nan")
    with pytest.raises(QuizPackError, match="finite"):
        validate_pack(pack)


def test_telemetry_rejects_phi_and_free_text_payloads(tmp_path: Path):
    telemetry = QuizTelemetry(tmp_path)
    with pytest.raises(ValueError, match="payload field"):
        telemetry.record(
            "pack_started",
            pack_id="pack-v1",
            case_id="35",
            mode="solo",
            payload={"participant_name": "Patient Name"},
        )
    with pytest.raises(ValueError, match="category"):
        telemetry.record(
            "content_report",
            pack_id="pack-v1",
            case_id="35",
            mode="solo",
            payload={"category": "free-form clinical note"},
        )
    assert not telemetry.path.exists()


def test_registry_is_immutable_and_seeded_playlist_is_non_repeating_and_balanced():
    registry = QuizPackRegistry(catalog_with())
    first = registry.get("test-pack-0-v1")
    first["title"] = "mutated"
    assert registry.get("test-pack-0-v1")["title"] != "mutated"
    selected = []
    for _ in range(10):
        selected.append(registry.select(
            "mixed-challenge-v1",
            seed="one-session",
            exclude_pack_ids=selected,
        )["pack_id"])
    statuses = [registry.get(pack_id)["ground_truth"]["lesion_present"] for pack_id in selected]
    assert statuses.count(True) == 7
    assert statuses.count(False) == 3
    excluded = selected[:]
    with pytest.raises(QuizPackError, match="non-repeating"):
        registry.select("mixed-challenge-v1", seed="again", exclude_pack_ids=excluded)


def test_unreviewed_registry_mode_exposes_pending_packs_only_when_explicitly_enabled():
    catalog = catalog_with(1)
    catalog["packs"][0]["approval"] = {"status": "pending"}
    default_registry = QuizPackRegistry(catalog)
    with pytest.raises(QuizPackError, match="not approved"):
        default_registry.get("test-pack-0-v1")
    development_registry = QuizPackRegistry(catalog, allow_unreviewed=True)
    assert development_registry.get("test-pack-0-v1")["approval"]["status"] == "pending"
    assert development_registry.public_playlists()[0]["pack_count"] == 1


def test_generator_positive_normal_multiple_lesion_and_failures(tmp_path: Path):
    root = tmp_path / "dataset"
    write_case(root, 1, lesion=True, multiple=True)
    write_case(root, 2, lesion=False)
    write_case(root, 3, lesion=True, mask_affine=np.diag([2, 1, 1, 1]))
    profiles = profile_for(1, 2, 3)
    ledger = {"approvals": {}}
    positive = generate_candidate(
        root,
        Candidate("1", "PanTS_00000001", True, (1, 1, 1), {
            "structured report": synthetic_structured_report(lesion=True),
        }),
        profiles,
        ledger,
        qa_dir=tmp_path / "qa",
    )
    normal = generate_candidate(root, Candidate("2", "PanTS_00000002", False, (1, 1, 1), {
        "structured report": synthetic_structured_report(lesion=False),
    }), profiles, ledger)
    assert positive.pack["ground_truth"]["lesion_present"] is True
    assert positive.pack["ground_truth"]["pancreatic_region"] in {"head", "body", "tail"}
    assert positive.pack["ground_truth"]["location_distance_mm"] >= 0
    assert positive.pack["pack_id"] == "pants-case-00000001-v2"
    assert [item["id"] for item in positive.pack["questions"]] == [
        "organ", "presence", "location", "report_finding", "conclusion"
    ]
    assert any(item["code"] == "multiple_lesions" for item in positive.warnings)
    assert positive.qa_image
    qa = Image.open(positive.qa_image).convert("RGB")
    assert np.any(np.all(np.asarray(qa) == np.array([80, 180, 255]), axis=2))
    quarantined = generate_candidate(
        root,
        Candidate("1", "PanTS_00000001", True, (1, 1, 1), {}),
        profiles,
        {
            "approvals": {"pants-case-00000001-v2": {"status": "approved"}},
            "reviewed_cohorts": ["pancreas-imaging-chain/2.0.0:test-profile"],
            "quarantined_cohorts": ["pancreas-imaging-chain/2.0.0:test-profile"],
        },
    )
    assert quarantined.pack["approval"]["status"] == "quarantined"
    assert normal.pack["questions"][1]["correct_choice_id"] == "no"
    assert normal.pack["questions"][2]["correct_choice_id"] == "not_applicable"
    with pytest.raises(QuizGenerationError) as mismatch:
        generate_candidate(root, Candidate("3", "PanTS_00000003", True, None, {}), profiles, ledger)
    assert mismatch.value.code == "mismatched_geometry"
    write_case(root, 7, lesion=True, affine=np.diag([1.0, 1.0, 1e-12, 1.0]))
    with pytest.raises(QuizGenerationError) as singular:
        generate_candidate(root, Candidate("7", "PanTS_00000007", True, None, {}), profile_for(7), ledger)
    assert singular.value.code == "mismatched_geometry"
    with pytest.raises(QuizGenerationError) as missing:
        generate_candidate(root, Candidate("4", "PanTS_00000004", True, None, {}), profiles, ledger)
    assert missing.value.code == "missing_ct"
    image_dir = root / "image_only" / "PanTS_00000006"
    image_dir.mkdir(parents=True)
    nib.save(nib.Nifti1Image(np.zeros((4, 4, 2)), np.eye(4)), image_dir / "ct.nii.gz")
    with pytest.raises(QuizGenerationError) as missing_mask:
        generate_candidate(root, Candidate("6", "PanTS_00000006", True, None, {}), profile_for(6), ledger)
    assert missing_mask.value.code == "missing_mask"
    write_case(root, 5, lesion=True)
    with pytest.raises(QuizGenerationError) as unknown:
        generate_candidate(root, Candidate("5", "PanTS_00000005", True, None, {}), profiles, ledger)
    assert unknown.value.code == "unknown_label_profile"


def test_discover_candidates_reads_positivity_from_validated_combined_labels(tmp_path: Path):
    root = tmp_path / "dataset"
    write_case(root, 1, lesion=True)
    write_case(root, 2, lesion=False)
    candidates = discover_candidates(root, profile_for(1, 2))
    assert [(item.case_id, item.tumor_positive) for item in candidates] == [("1", True), ("2", False)]
    assert all(item.metadata["candidate_source"] == "combined_labels" for item in candidates)


def test_size_outliers_are_flagged_for_review():
    generated = []
    for index, size in enumerate((10.0, 11.0, 12.0, 13.0, 100.0)):
        pack = copy.deepcopy(get_pack("radworld-case-35-v1"))
        pack["pack_id"] = f"outlier-{index}"
        pack["ground_truth"]["max_axial_diameter_mm"] = size
        pack["ground_truth"]["location_distance_mm"] = size
        generated.append(GeneratedCandidate(pack, []))
    _mark_outliers(generated)
    assert not generated[2].warnings
    assert any(item["code"] == "size_outlier" for item in generated[-1].warnings)
    assert any(item["code"] == "location_outlier" for item in generated[-1].warnings)


def test_review_sample_always_includes_warning_cases():
    generated = []
    for index, pack in enumerate(catalog_with(100)["packs"]):
        generated.append(GeneratedCandidate(pack, [
            {"code": "multiple_lesions", "message": "review required"}
        ] if index == 99 else []))
    selected = _review_sample(generated, 100, "warning-sample", {
        "reviewed_template_versions": ["pancreas-imaging-chain/1.1.0"],
        "reviewed_label_profiles": ["case35-reviewed-v1"],
    })
    assert len(selected) >= 20
    assert "test-pack-99-v1" in selected
    assert len(_review_sample(generated, 100, "new-cohort", {})) == 100


def test_room_freezes_pack_and_solo_uses_identical_scoring_and_reveal(tmp_path: Path):
    pants = tmp_path / "pants"
    write_case(pants, 35, lesion=True)
    base = copy.deepcopy(get_pack("radworld-case-35-v1"))
    base["ground_truth"]["crosshair_lps"] = [-3.5, -11.0, 2.0]
    base["ground_truth"]["lesion_centroid_lps"] = [-3.5, -11.0, 2.0]
    base["approval"] = {"status": "approved", "reviewer": "test-fixture"}
    for question in base["questions"]:
        question["viewer_cue"]["crosshair_lps"] = base["ground_truth"]["crosshair_lps"]
    base["questions"][-1]["reveal_viewer_cue"].update({
        "crosshair_lps": base["ground_truth"]["crosshair_lps"],
        "reference_measurement_lps": base["ground_truth"]["reference_measurement_lps"],
    })
    registry = QuizPackRegistry({"catalog_id": "freeze-v1", "catalog_version": 1, "packs": [base]})
    room_store = LiveRoomStore(tmp_path / "sessions", pants, quiz_registry=registry)
    metadata, key, host_secret = room_store.create_quiz_room("35")
    with pytest.raises(LiveRoomError, match="case_id must be numeric"):
        room_store.create_quiz_room("not-a-case")
    frozen_path = room_store.root / metadata["room_id"] / "quiz_pack.json"
    assert frozen_path.is_file()
    frozen = json.loads(frozen_path.read_text())
    base["questions"][0]["correct_choice_id"] = "liver"
    assert json.loads(frozen_path.read_text()) == frozen

    practice = QuizPracticeStore(tmp_path / "sessions", pants, registry=registry)
    attempt, attempt_key = practice.start(base["pack_id"])
    answers = {
        question["id"]: question["correct_choice_id"]
        for question in frozen["questions"]
    }
    result = practice.submit(attempt["attempt_id"], attempt_key, answers)
    assert result["score"] == 4
    assert result["reveals"] == [
        reveal_for(frozen, question["id"], {"solo": {"choice_id": answers[question["id"]]}})
        for question in frozen["questions"]
    ]
    assert practice.reveal_segmentation(attempt["attempt_id"], attempt_key)
    with pytest.raises(QuizPracticeError, match="already submitted"):
        practice.submit(attempt["attempt_id"], attempt_key, answers)


def test_quiz_submit_is_single_writer_and_expired_attempts_are_cleaned(tmp_path: Path):
    pants = tmp_path / "pants"
    write_case(pants, 100, lesion=True)
    pack = catalog_with(1)["packs"][0]
    registry = QuizPackRegistry({"catalog_id": "concurrency-v1", "catalog_version": 1, "packs": [pack]})
    practice = QuizPracticeStore(tmp_path / "sessions", pants, registry=registry)
    attempt, attempt_key = practice.start(pack["pack_id"])
    answers = {question["id"]: question["correct_choice_id"] for question in pack["questions"]}
    barrier = threading.Barrier(2)
    outcomes: list[str] = []

    def submit() -> None:
        barrier.wait()
        try:
            practice.submit(attempt["attempt_id"], attempt_key, answers)
            outcomes.append("accepted")
        except QuizPracticeError:
            outcomes.append("rejected")

    workers = [threading.Thread(target=submit) for _ in range(2)]
    for worker in workers:
        worker.start()
    for worker in workers:
        worker.join()
    assert sorted(outcomes) == ["accepted", "rejected"]

    attempt_path = practice.root / attempt["attempt_id"] / "attempt.json"
    expired = json.loads(attempt_path.read_text())
    expired["expires_at"] = "2000-01-01T00:00:00Z"
    attempt_path.write_text(json.dumps(expired))
    assert practice.cleanup_expired() == [attempt["attempt_id"]]
    assert not attempt_path.parent.exists()


@pytest.mark.parametrize(("with_report_question", "expected_count"), [(True, 5), (False, 4)])
def test_v2_pack_completes_identically_in_solo_and_race(
    tmp_path: Path, with_report_question: bool, expected_count: int
):
    pants = tmp_path / "pants"
    write_case(pants, 100, lesion=True)
    pack = v2_pack(catalog_with(1)["packs"][0], with_report_question=with_report_question)
    registry = QuizPackRegistry({"catalog_id": "non-case35-v2", "catalog_version": 1, "packs": [pack]})
    correct_answers = {
        question["id"]: question["correct_choice_id"]
        for question in pack["questions"]
    }

    practice = QuizPracticeStore(tmp_path / "sessions", pants, registry=registry)
    attempt, attempt_key = practice.start(pack["pack_id"])
    solo = practice.submit(attempt["attempt_id"], attempt_key, correct_answers)
    assert attempt["pack"]["case_id"] == "100"
    assert solo["score"] == expected_count
    assert solo["max_score"] == expected_count

    rooms = LiveRoomStore(tmp_path / "sessions", pants, quiz_registry=registry)
    metadata, room_key, host_secret = rooms.create_quiz_room(
        "100", quiz_pack_id=pack["pack_id"], quiz_timer_seconds=None
    )
    room_id = metadata["room_id"]
    rooms.connect_quiz_host(room_id, room_key, host_secret, "host")
    state, _ = rooms.start_quiz(
        room_id,
        room_key,
        host_secret,
        "host",
        [{"participant_id": "student", "name": "Student"}],
    )
    race_reveals = []
    for index, question in enumerate(pack["questions"]):
        assert state["current_question"]["id"] == attempt["pack"]["questions"][index]["id"]
        assert {
            choice["id"] for choice in state["current_question"]["choices"]
        } == {
            choice["id"] for choice in attempt["pack"]["questions"][index]["choices"]
        }
        state, _, closed = rooms.answer_quiz(
            room_id,
            room_key,
            "student",
            "Student",
            correct_answers[question["id"]],
            {"student"},
        )
        assert closed is not None and state["phase"] == "question_closed"
        state, _ = rooms.reveal_quiz_question(room_id, room_key, host_secret, "host")
        race_reveals.append(state["reveal"])
        state, _ = rooms.advance_quiz(
            room_id,
            room_key,
            host_secret,
            "host",
            [{"participant_id": "student", "name": "Student"}],
        )
    assert state["phase"] == "completed"
    assert len(race_reveals) == expected_count
    assert race_reveals == solo["reveals"]
    leaderboard = {row["participant_id"]: row for row in state["leaderboard"]}
    assert leaderboard["student"]["score"] == expected_count
    assert leaderboard["student"]["max_score"] == expected_count
    if with_report_question:
        assert race_reveals[3]["source_label"] == "Structured report finding"
    with zipfile.ZipFile(rooms.build_export(room_id, room_key)) as archive:
        summary = json.loads(archive.read("quiz-summary.json"))
        exported = {row["participant_id"]: row for row in summary["leaderboard"]}
        assert exported["student"]["score"] == expected_count
        assert exported["student"]["max_score"] == expected_count
    race_mask = nib.Nifti1Image.from_bytes(gzip.decompress(
        rooms.get_quiz_reveal_segmentation(room_id, room_key)
    ))
    solo_mask = nib.Nifti1Image.from_bytes(gzip.decompress(
        practice.reveal_segmentation(attempt["attempt_id"], attempt_key)
    ))
    assert np.array_equal(np.asanyarray(race_mask.dataobj), np.asanyarray(solo_mask.dataobj))
    assert np.allclose(race_mask.affine, solo_mask.affine)


def test_playlist_and_solo_rest_payloads_keep_answers_private_until_submit(tmp_path: Path, monkeypatch):
    pants = tmp_path / "pants"
    write_case(pants, 35, lesion=True)
    practice = QuizPracticeStore(tmp_path / "sessions", pants)
    monkeypatch.setattr(education_api, "_quiz_practice_store", practice)
    app = Flask(__name__)
    app.register_blueprint(education_api.education_blueprint, url_prefix="/api")
    client = app.test_client()

    playlists = client.get("/api/education/quiz-playlists")
    assert playlists.status_code == 200
    selected = client.post(
        "/api/education/quiz-playlists/mixed-challenge-v1/select",
        json={"seed": "rest-test", "exclude_pack_ids": []},
    )
    assert selected.status_code == 200
    assert "correct_choice_id" not in selected.get_data(as_text=True)

    started = client.post("/api/education/quiz-packs/radworld-case-35-v1/attempts")
    assert started.status_code == 201
    started_body = started.get_json()
    assert "correct_choice_id" not in started.get_data(as_text=True)
    answers = {
        "organ": "pancreas",
        "presence": "yes",
        "location": "tail",
        "conclusion": "focal_tail_23mm",
    }
    completed = client.post(
        f"/api/education/quiz-attempts/{started_body['attempt_id']}/submit",
        headers={"X-Quiz-Attempt-Key": started_body["attempt_key"]},
        json={"answers": answers},
    )
    assert completed.status_code == 200
    assert completed.get_json()["score"] == 4
    assert "correct_choice_id" in completed.get_data(as_text=True)


@pytest.mark.parametrize("target", [50, 100, 500])
def test_full_synthetic_execution_reaches_exact_release_gates(tmp_path: Path, target: int):
    root = tmp_path / "dataset"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["PanTS ID", "spacing", "tumor?", "structured report"])
    approvals = {}
    target_positive, target_normal = TARGETS[target]
    case_ids = [*range(1, target + 1)] if target != 500 else [*range(1, 500), 9901]
    for index, case_id in enumerate(case_ids, start=1):
        positive = index <= target_positive
        warning_case = target == 50 and case_id == 1
        write_case(root, case_id, lesion=positive, multiple=warning_case)
        report_text = None if case_id == 9901 else synthetic_structured_report(lesion=positive)
        sheet.append([f"PanTS_{case_id:08d}", "(1.0, 1.0, 1.0)", int(positive), report_text])
        pack_id = "radworld-case-35-v2" if case_id == 35 else f"pants-case-{case_id:08d}-v2"
        approvals[pack_id] = {"status": "approved", "reviewer": "synthetic-test"}
        if warning_case:
            approvals[pack_id]["acknowledged_warnings"] = [
                "multiple_lesions", "location_outlier"
            ]
        if case_id == 9901:
            approvals[pack_id]["acknowledged_warnings"] = ["report_missing"]
    metadata = tmp_path / "metadata.xlsx"
    workbook.save(metadata)
    profiles = tmp_path / "profiles.json"
    profiles.write_text(json.dumps({
        "profile_version": 1,
        "profiles": [{
            "profile_id": "synthetic-profile",
            "match": {"case_ranges": [[1, target]], "case_ids": [9901] if target == 500 else []},
            "labels": {
                "pancreas": [17, 18, 19, 20],
                "head": [19],
                "body": [18],
                "tail": [20],
                "lesion": [28],
            },
            "reveal_output_label": 1,
            "mesh_organ_id": 22,
        }],
    }))
    ledger = tmp_path / "ledger.json"
    ledger_data = {
        "ledger_version": 1,
        "approvals": approvals,
        "difficulty_overrides": {},
        "reviewed_template_versions": ["pancreas-imaging-chain/2.0.0"],
        "reviewed_label_profiles": ["synthetic-profile"],
        "reviewed_cohorts": ["pancreas-imaging-chain/2.0.0:synthetic-profile"],
        "quarantined_cohorts": [],
    }
    ledger.write_text(json.dumps(ledger_data))
    catalog = tmp_path / "catalog.json"
    draft_catalog = tmp_path / "draft-catalog.json"
    report_path = tmp_path / "report.json"
    if target == 50:
        blocked_ledger = tmp_path / "blocked-ledger.json"
        blocked_data = copy.deepcopy(ledger_data)
        blocked_data["approvals"]["pants-case-00000001-v2"].pop("acknowledged_warnings")
        blocked_ledger.write_text(json.dumps(blocked_data))
        blocked_catalog = tmp_path / "blocked-catalog.json"
        blocked = generate_release(
            dataset_root=root,
            metadata_path=metadata,
            target=50,
            output_catalog=blocked_catalog,
            report_path=tmp_path / "blocked-report.json",
            qa_dir=None,
            profile_path=profiles,
            ledger_path=blocked_ledger,
            seed="synthetic-50",
        )
        assert blocked["acceptance"]["passed"] is False
        assert blocked["acceptance"]["review_required_pack_count"] == 1
        assert not blocked_catalog.exists()
        pending_ledger = tmp_path / "pending-ledger.json"
        pending_data = copy.deepcopy(ledger_data)
        pending_data["approvals"]["pants-case-00000002-v2"]["status"] = "pending"
        pending_ledger.write_text(json.dumps(pending_data))
        pending = generate_release(
            dataset_root=root,
            metadata_path=metadata,
            target=50,
            output_catalog=tmp_path / "pending-catalog.json",
            report_path=tmp_path / "pending-report.json",
            qa_dir=None,
            profile_path=profiles,
            ledger_path=pending_ledger,
            seed="synthetic-50",
        )
        assert pending["acceptance"]["review_required_pack_count"] == 1
    report = generate_release(
        dataset_root=root,
        metadata_path=metadata,
        target=target,
        output_catalog=catalog,
        report_path=report_path,
        qa_dir=None,
        draft_catalog_path=draft_catalog,
        profile_path=profiles,
        ledger_path=ledger,
        seed=f"synthetic-{target}",
    )
    assert report["acceptance"]["passed"] is True
    assert report["acceptance"]["review_required_pack_count"] == 0
    assert report["acceptance"]["composition"] == {
        "positive": target_positive,
        "normal": target_normal,
    }
    assert report["acceptance"]["report_question_count"] == target - (1 if target == 500 else 0)
    assert report["acceptance"]["missing_report_pack_count"] == (1 if target == 500 else 0)
    assert report["report_enrichment"]["question_family_counts"]["conclusion"] == target
    assert len(report["review_sample_pack_ids"]) == (target if target == 50 else target // 5)
    published = json.loads(catalog.read_text())
    assert len(published["packs"]) == target
    case35 = next(item for item in published["packs"] if item["case_id"] == "35")
    assert case35["pack_id"] == "radworld-case-35-v2"
    QuizPackRegistry(published)
    if target == 500:
        missing_pack = next(item for item in published["packs"] if item["case_id"] == "9901")
        assert missing_pack["pack_id"] == "pants-case-00009901-v2"
        assert [item["id"] for item in missing_pack["questions"]] == [
            "organ", "presence", "location", "conclusion"
        ]
        repeated_catalog = tmp_path / "catalog-repeat.json"
        repeated = generate_release(
            dataset_root=root,
            metadata_path=metadata,
            target=target,
            output_catalog=repeated_catalog,
            report_path=tmp_path / "report-repeat.json",
            qa_dir=None,
            profile_path=profiles,
            ledger_path=ledger,
            seed=f"synthetic-{target}",
        )
        assert repeated["acceptance"]["catalog_content_sha256"] == report["acceptance"]["catalog_content_sha256"]
        assert repeated_catalog.read_bytes() == catalog.read_bytes()
    draft = json.loads(draft_catalog.read_text())
    assert draft["review_only"] is True
    assert len(draft["packs"]) == target


def test_development_release_publishes_pending_structurally_valid_packs(tmp_path: Path):
    root = tmp_path / "dataset"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["PanTS ID", "spacing", "tumor?", "structured report"])
    for case_id in range(1, 51):
        positive = case_id <= 35
        write_case(root, case_id, lesion=positive)
        sheet.append([
            f"PanTS_{case_id:08d}", "(1.0, 1.0, 1.0)", int(positive),
            synthetic_structured_report(lesion=positive),
        ])
    metadata = tmp_path / "metadata.xlsx"
    workbook.save(metadata)
    profiles = tmp_path / "profiles.json"
    profiles.write_text(json.dumps({
        "profile_version": 1,
        "profiles": [{
            "profile_id": "development-profile",
            "match": {"case_ranges": [[1, 50]]},
            "labels": {"pancreas": [17, 18, 19, 20], "head": [19], "body": [18], "tail": [20], "lesion": [28]},
            "reveal_output_label": 1,
            "mesh_organ_id": 22,
        }],
    }))
    ledger = tmp_path / "ledger.json"
    ledger.write_text(json.dumps({"ledger_version": 1, "approvals": {}}))
    catalog_path = tmp_path / "catalog.json"
    report = generate_release(
        dataset_root=root,
        metadata_path=metadata,
        target=50,
        output_catalog=catalog_path,
        report_path=tmp_path / "report.json",
        qa_dir=None,
        profile_path=profiles,
        ledger_path=ledger,
        allow_unreviewed=True,
    )
    assert report["acceptance"]["passed"] is True
    assert report["acceptance"]["approved_pack_count"] == 0
    assert report["acceptance"]["published_pack_count"] == 50
    catalog = json.loads(catalog_path.read_text())
    assert catalog["release_stage"] == "development"
    assert catalog["review_only"] is True
    assert len(catalog["packs"]) == 50
    assert all(item["version"] == 2 and item["approval"]["status"] == "pending" for item in catalog["packs"])
    assert all(len(item["questions"]) == 5 for item in catalog["packs"])
    assert QuizPackRegistry(catalog, allow_unreviewed=True).public_playlists()[2]["pack_count"] == 50


def test_local_artifacts_only_ignores_unstaged_metadata_candidates(tmp_path: Path):
    root = tmp_path / "dataset"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["PanTS ID", "spacing", "tumor?", "structured report"])
    for case_id in range(1, 52):
        positive = case_id <= 35
        if case_id != 51:
            write_case(root, case_id, lesion=positive)
        sheet.append([
            f"PanTS_{case_id:08d}", "(1.0, 1.0, 1.0)", int(positive),
            synthetic_structured_report(lesion=positive),
        ])
    metadata = tmp_path / "metadata.xlsx"
    workbook.save(metadata)
    profiles = tmp_path / "profiles.json"
    profiles.write_text(json.dumps({
        "profile_version": 1,
        "profiles": [{
            "profile_id": "local-profile",
            "match": {"case_ranges": [[1, 51]]},
            "labels": {"pancreas": [17, 18, 19, 20], "head": [19], "body": [18], "tail": [20], "lesion": [28]},
        }],
    }))
    ledger = tmp_path / "ledger.json"
    ledger.write_text(json.dumps({"ledger_version": 1, "approvals": {}}))
    report = generate_release(
        dataset_root=root,
        metadata_path=metadata,
        target=50,
        output_catalog=tmp_path / "catalog.json",
        report_path=tmp_path / "report.json",
        qa_dir=None,
        profile_path=profiles,
        ledger_path=ledger,
        allow_unreviewed=True,
        local_artifacts_only=True,
    )
    assert report["run"]["candidate_source"] == "local_metadata_artifacts"
    assert report["run"]["attempted_candidates"] == 50
    assert report["acceptance"]["passed"] is True


def test_official_mask_archive_extraction_is_safe_atomic_and_resumable(tmp_path: Path):
    archive_path = tmp_path / "labels.tar.gz"
    with tarfile.open(archive_path, mode="w:gz") as archive:
        for name, payload in (
            ("PanTS_00000035/combined_labels.nii.gz", b"valid-mask"),
            ("../escape/combined_labels.nii.gz", b"unsafe"),
            ("PanTS_00000035/segmentations/pancreas.nii.gz", b"unused"),
        ):
            member = tarfile.TarInfo(name)
            member.size = len(payload)
            archive.addfile(member, io.BytesIO(payload))

    dataset_root = tmp_path / "dataset"
    first = extract_combined_labels(archive_path, dataset_root)
    staged = dataset_root / "mask_only" / "PanTS_00000035" / "combined_labels.nii.gz"
    assert first == {"extracted": 1, "existing": 0}
    assert staged.read_bytes() == b"valid-mask"
    assert not (tmp_path / "escape").exists()
    assert extract_combined_labels(archive_path, dataset_root) == {"extracted": 0, "existing": 1}
